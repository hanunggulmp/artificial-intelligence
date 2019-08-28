import xlrd
import openpyxl
import math
import operator
from operator import itemgetter

wb = xlrd.open_workbook('Dataset Tugas 3 AI 1718.xlsx')

worksheetRead = wb.sheet_by_index(0)  # read first page of sheet based on index
datasheet = wb.sheet_by_index(1)  # read second page of sheet based on index

table = list()
record = list()
dataTestTable = list()

def createTable(sheet): #Create a Table for each sheet
    global record
    listTable = list()
    for x in range(1, sheet.nrows):
        for y in range(1, sheet.ncols):
            record.append(sheet.cell(x, y).value)
        listTable.append(record)
        record = []
        x += 1
    return listTable

k_fold=list()
data1=list()
data2=list()
data3=list()
data4=list()
data5=list()
data6=list()
data7=list()
data8=list()

for a in range(worksheetRead.nrows-4000, worksheetRead.nrows-3651):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data1.append(record)
    record = []
for a in range(worksheetRead.nrows-3500, worksheetRead.nrows- 3151):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data2.append(record)
    record = []
for a in range(worksheetRead.nrows-3000, worksheetRead.nrows-2651):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data3.append(record)
    record = []
for a in range(worksheetRead.nrows-2500, worksheetRead.nrows-2151):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data4.append(record)
    record = []
for a in range(worksheetRead.nrows-2000, worksheetRead.nrows-1651):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data5.append(record)
    record = []
for a in range(worksheetRead.nrows-1500, worksheetRead.nrows-1151):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data6.append(record)
    record = []
for a in range(worksheetRead.nrows-1000, worksheetRead.nrows-651):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data7.append(record)
    record = []
for a in range(worksheetRead.nrows-500, worksheetRead.nrows - 151):
    for i in range(1, worksheetRead.ncols):
        record.append(worksheetRead.cell(a,i).value)
    data8.append(record)
    record = []

k_fold = [data1, data2, data3, data4, data5, data6, data7, data8]
length = len(k_fold) -1
sk = list()
K = 0
acc = float
avg = list()
foldtest = 0

def k_foldTester(kMin, kMax): #Nge Test K-fold buat cari nilai K
    global foldtest
    global K
    global avg
    while (kMin<=kMax):
        temp = []
        foldtest = 1
        for a in range(length):
            dSet = k_fold[a]
            dTrain =[]
            b = 0
            sk = []
            while (b<=length):
                if b == a:
                    b+=1
                else:
                    dTrain.append(k_fold[b])
                    n = k_fold[b]
                    for r in range(len(k_fold[b])):
                        sk.append(n[r])
                b += 1
            test_K(dSet,sk,kMin)
            foldtest += 1
            temp.append(acc)
        avgtemp = sum(temp)/(len(temp))
        print("avg for K",kMin,"is",avgtemp)
        avg.append((kMin,avgtemp))
        avg.sort(key=itemgetter(1),reverse=True)
        print("average = ",avg)
        kMin+=2
    K = avg[0][0]

def euclideanDist(x, x1): #Cari Nilai Euclide
    d =0.0
    for i in range(len(x)-1):
        d +=pow (float(x[i]) - float(x1[i]),2)
    d = math.sqrt(d)
    return d

def test_K(ts, tr, k_value): #nge test hasil class berdasarkan k tertentu,
    global acc
    order = 1
    global foldtest
    for i in ts:
        eu_Distance = []
        knn = []
        s = []
        hoax = 0
        truth = 0
        for j in tr:
            euval = euclideanDist(i, j)
            eu_Distance.append((j[4], euval))
            eu_Distance.sort(key=operator.itemgetter(1))
            knn = eu_Distance[:k_value]
            print(k_value,"fold ke",foldtest)
            print("data ke", order)
        for k in knn:
            if k[0] == 1:
                hoax += 1
            else:
                truth += 1
        print("hoax is", hoax)
        print("not hoax is", truth)
        if hoax > truth:
            i.append(1.0)
        elif hoax < truth:
            i.append(0.0)
        else:
            i.append('error')
        print("\n")
        order += 1
    acc = countAccuracy(ts)
    for s in ts:
        del s[5]
        print("i value = ",s)


def countAccuracy(ts): #ngitung akurasi berdasrakan fungsi test_k
    trueValue = 0
    for x in ts:
        if x[4] == x[5]:
            trueValue += 1
    accuracy = float(trueValue) / len(ts) * 100
    return accuracy



def knn_test(ktest,test,train): #nulis jawaban di lembar sheet excel
    row = 2
    col = 6
    for i in test:
        eu_Distance = []
        knn = []
        hoax = 0
        truth = 0
        for j in train:
            euval = euclideanDist(i, j)
            eu_Distance.append((j[4], euval))
            eu_Distance.sort(key=operator.itemgetter(1))
            knn = eu_Distance[:ktest]
            print("ROW",row)
        for k in knn:
            if k[0] == 1:
                hoax += 1
            else:
                truth += 1
        print("hoax is", hoax)
        print("not hoax is", truth)
        if hoax > truth:
            i.append(1.0)
            sheet.cell(row = row,column = col).value = 1.0
        elif hoax < truth:
            i.append(0.0)
            sheet.cell(row = row, column = col).value = 0.0
        else:
            i.append('null')
        row += 1

#---------------------MAIN-------------------#
dataworkWrite = openpyxl.load_workbook('Dataset Tugas 3 AI 1718.xlsx')
sheet = dataworkWrite.get_sheet_by_name('DataTest')

table = createTable(worksheetRead)
dataTestTable = createTable(datasheet)

dataTemp = list()
tab = 0
dataTrain = []

while (tab <= length):
    dataTemp.append(k_fold[tab])
    n = k_fold[tab]
    for r in range(len(k_fold[tab])):
        dataTrain.append(n[r])
    tab += 1

k_foldTester(13,83)
knn_test(K,dataTestTable,dataTrain)

dataworkWrite.save("Final Result Tugas 3 AI Fixed.xlsx")
print("the best K is",K,"for acc is:",avg)
print("length",len(dataTestTable),"the data table is",dataTestTable)
print("table",table)